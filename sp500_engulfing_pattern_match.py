# Finds engulfing pattern trades based on conditions and savees the data. The data can be used for evaluation of a trading
# strategy

import csv
import datetime
import openpyxl
import os
import yfinance as yf

start = datetime.datetime.now()

def is_bearish_candle(candle):
    return float(candle['Close']) < float(candle['Open'])

def is_bullish_candle(candle):
    return float(candle['Close']) > float(candle['Open'])

def is_bullish_engulfing(candles, index):
    current_day = candles[index]
    previous_day = candles[index-1]

    return is_bearish_candle(previous_day) \
        and float(current_day['Open']) < float(previous_day['Close']) \
        and float(current_day['Close']) > float(previous_day['Open'])

def is_bearish_engulfing(candles, index):
    current_day = candles[index]
    previous_day = candles[index-1]

    return is_bullish_candle(previous_day) \
        and float(current_day['Open']) > float(previous_day['Close']) \
        and float(current_day['Close']) < float(previous_day['Open'])

def avg_trend(candles, index, num_trend_days): # %
    prev_day_price = float(candles[index - 1]['Close'])
    sum_diff = 0    # start price - day price

    for i in range (index - num_trend_days, index-2):
        sum_diff += (prev_day_price - float(candles[i]['Close'])) / (index-i)
    return (sum_diff / (num_trend_days - 2)) / prev_day_price * 100

def abs_trend(candles, index, num_trend_days): # %
    day1 = float(candles[index - num_trend_days]['Close'])
    prev_day = float(candles[index - 1]['Close'])
    return (prev_day - day1) / day1 * 100

def candle_size(candle): # %
    return abs(float(candle['Close']) - float(candle['Open'])) / float(candle['Open']) * 100

def purchase(candles, index):
    current_day = candles[i]
    previous_day = candles[i-1]
    
    prev_day_mid_point = (float(previous_day['Open']) + float(previous_day['Close'])) / 2
    if float(current_day['Open']) * multiplier >= prev_day_mid_point * multiplier:
        return float(curr_candle['Open'])
    elif float(current_day['High']) * multiplier >= prev_day_mid_point * multiplier:
        return prev_day_mid_point
    else:
        return False

def profit():
    return round((sell_price - position['Buy']) / position['Buy'] * 100 * multiplier, 4)

def profit_target(price, target):
    return round(price * (1 + multiplier * (target/100)), 4)

def stop_loss(price):
    return round(price * (1 - multiplier * (stoploss/100)), 4)    

# Create summary workbook
summary_wb = openpyxl.Workbook()
summary_sheet = summary_wb['Sheet']
summary_wb_row = 2 # Add 1 at end of company iteration

# Write summary labels to summary_wb
summary_labels = ('Filename', 'Engulfing', 'Trend Days', 'Trend Rate', 'Prev. Day Candle %', \
                'Curr. Day Candle %', 'Stoploss', 'Hold Days', 'Number of Trades', \
                'Invested Capital', 'Total Profit',	'Profit/Trade $', 'Profit/Trade %', \
                'Std. Dev. %', 'Total Losses', 'Loss %')
for col in range(1, len(summary_labels) + 1):
    summary_sheet.cell(1, col).value = summary_labels[col-1]
del summary_labels

summary_sheet.freeze_panes = 'A2'

# Initialise
count = 1
profit_target_pc = 3 # Can iterate through this later.
criteria = ('Engulfing', 'Trend Days', 'Trend Rate', 'Prev. Day Candle %', \
                                'Cur. Day Candle %', 'Stoploss', 'Hold Days')

# Iterate through different conditions
print('Iterating through conditions...')
for engulfing in ('bullish', 'bearish'):
    for trend_days in (10,): #(3, 5, 7, 10):
        for trend_rate in (0.5, 1, 2, 3, 5):
            for yesterday_candle_size in range(5):
                for today_candle_size in range(5):
                    if today_candle_size < yesterday_candle_size:
                        continue
                    for stoploss in (2, 3, 5, 10, 15, 20):
                        for hold_days in (5,  6, 10, 14):
                            
                            # Create new workbook
                            wb = openpyxl.Workbook()
                            sheet = wb['Sheet']
                            wb_row = 2 # First row in wb to enter data

                            # Print criteria          
                            criteria_values = (engulfing, trend_days, trend_rate, yesterday_candle_size, \
                                            today_candle_size, stoploss, hold_days)

                            # Write criteria labels and values to wb
                            for row_num in range(2, len(criteria) + 2):
                                sheet[f'L{row_num}'].value = criteria[row_num - 2]
                                sheet[f'M{row_num}'].value = criteria_values[row_num - 2]

                            # Print outcome labels to workbook
                            outcomes = ('Number of Trades', 'Invested Capital', 'Total Profit', 'Profit/Trade $', \
                                'Profit/Trade %', 'Std. Dev. %', 'Total Losses', 'Loss %')
                            for row_num in range(2, len(outcomes) + 2):
                                sheet[f'N{row_num}'].value = outcomes[row_num - 2]

                            # Print trade info
                            trade_info = ('Date', 'Year', 'Company', 'Ticker', 'Buy $', 'Sell $', 'Stoploss', 'Profit $', 'Profit %')
                            for col_num in range(1, len(trade_info) + 1):
                                sheet.cell(1, col_num).value = trade_info[col_num-1]

                            sheet.freeze_panes = 'A2'

                            if engulfing == 'bullish':
                                multiplier = 1
                                price_extreme_sl = 'Low'
                                price_extreme_pt = 'High'
                            else:
                                multiplier = -1
                                price_extreme_sl = 'High'
                                price_extreme_pt = 'Low'


                            companies = csv.reader(open("sp500.csv"))

                            for company in companies:

                                ticker, company_name = company
                                history_file = open('history_edit/{}.csv'.format(ticker))
                                reader = csv.DictReader(history_file)
                                candles = list(reader)

                                potential_buy = False

                                held_positions = []

                                # Iterate through candles. Check if condition is met.
                                for i in range(trend_days, len(candles)):
                                    curr_candle = candles[i]
                                    prev_candle = candles[i-1]

                                    # Check if buy condition filled
                                    if potential_buy:
                                        price = purchase(candles, i)
                                        if price:
                                            held_positions += [{'Date': curr_candle['Date'], 'Index': i, 
                                                            'Buy': price, 'Stoploss': stop_loss(price),
                                                            'Engulfing': engulfing, 'Profit target': 
                                                            profit_target(price, profit_target_pc)}]
                                    
                                    # Look through held position for sells
                                    for position in held_positions[:]:
                                        # Stoploss and days held
                                        if (float(curr_candle['Open']) * multiplier <= position['Stoploss'] * multiplier) \
                                            or float(curr_candle[price_extreme_sl]) * multiplier <= position['Stoploss'] * multiplier \
                                            or (i - position['Index'] >= hold_days):
                                            
                                            if float(curr_candle[price_extreme_sl]) * multiplier <= position['Stoploss'] * multiplier:
                                                sell_price = float(position['Stoploss'])
                                            else:
                                                sell_price = float(curr_candle['Open'])
                                            
                                            held_positions.remove(position)

                                            year = position['Date'][:4]
                                            profit_dollar = profit()/100 * 1000
                                            profit_pc = profit()
                                            trade_data = (position['Date'], year, company_name, ticker, position['Buy'], \
                                                        sell_price, position['Stoploss'], profit_dollar, profit_pc)
                                            
                                            # Trade info to workbook
                                            for col_num in range(1, len(trade_data) + 1):
                                                sheet.cell(wb_row, col_num).value = trade_data[col_num -1]
                                            
                                            del trade_data
                                            
                                            wb_row += 1

                                        # look for profit target to move stoploss
                                        elif float(curr_candle['Open']) * multiplier >= float(position['Profit target']) * multiplier \
                                        or float(curr_candle[price_extreme_pt]) *  multiplier >= float(position['Profit target']) * multiplier:
                                            held_positions.remove(position)
                                            held_position = position.copy()
                                            # Change profit target
                                            held_position['Profit target'] = position['Buy'] * 2
                                            # Make stoploss the buy price
                                            held_position['Stoploss'] = position['Buy']
                                            held_positions.append(held_position)

                                    # Reinitialze potential_buy
                                    potential_buy = False

                                    # Check engulfing
                                    if engulfing == 'bullish':
                                        if not is_bullish_engulfing(candles, i):
                                            continue
                                    else:
                                        if not is_bearish_engulfing(candles, i):
                                            continue
                                    
                                    # Check trend rate and type
                                    if not avg_trend(candles, i, trend_days) * (-1 * multiplier) >= trend_rate:
                                        continue

                                    # Yesterday candle size
                                    if not candle_size(prev_candle) >= yesterday_candle_size:
                                        continue

                                    # Today candle size
                                    if not candle_size(curr_candle) >= today_candle_size:
                                        continue

                                    # Consider buy
                                    potential_buy = True

                            # Count trades  
                            max_row = 1
                            check_rows = len(outcomes) + 1 # max_row may not be determined by number of trades.

                            if sheet.max_row == check_rows:
                                for r in range(check_rows, 1, -1):
                                    if sheet.cell(r, 1).value == None:
                                        max_row = r
                                        break
                            else: 
                                max_row = sheet.max_row

                            if max_row > 1:
                                # Total trades
                                total_trades = max_row - 1
                                
                                # Invested Capital
                                invested_capital = total_trades * 1000

                                # Total_profit
                                total_profit = 0
                                for r in range(2, max_row + 1):
                                    total_profit += sheet[f'H{r}'].value
                                av_profit_per_trade = round(total_profit / total_trades, 4)
                                av_profit_per_trade_pc = round(total_profit / invested_capital * 100, 4)

                                var_x_total_trades = 0
                                total_losses = 0
                                
                                # Standard deviation and total losses
                                for r in range(2, max_row+1):
                                    profit_pc = sheet[f'I{r}'].value
                                    var_x_total_trades += (profit_pc - av_profit_per_trade_pc) ** 2
                                    if profit_pc < 0:
                                        total_losses += 1
                                std = (var_x_total_trades / (total_trades-1)) ** 0.5
                                
                                # % of losses
                                loss_pc = round(total_losses / total_trades * 100, 4)

                                del var_x_total_trades
                                del profit_pc

                                outcomes = (total_trades, invested_capital, total_profit, av_profit_per_trade, \
                                        av_profit_per_trade_pc, std, total_losses, loss_pc)
                                
                                # Write outcomes to workbook
                                for r in range(2, len(outcomes) + 2):
                                    sheet[f'O{r}'].value = outcomes[r-2]

                            xl_filename = (f'Trade data/{engulfing[:2]}-TD{trend_days}-TR{trend_rate}-'\
                                f'PCS{yesterday_candle_size}-CCS{today_candle_size}-SL{stoploss}'\
                                f'-HD{hold_days}.xlsx')

                            # Copy data to summary_wb
                            summary_data =(xl_filename, engulfing, trend_days, trend_rate, yesterday_candle_size, \
                                            today_candle_size, stoploss, hold_days, total_trades, invested_capital, \
                                            total_profit, av_profit_per_trade, av_profit_per_trade_pc, std, \
                                            total_losses, loss_pc)

                            for col in range(1, len(summary_data) + 1):
                                summary_sheet.cell(summary_wb_row, col).value = summary_data[col-1]
                            
                            summary_wb_row += 1

                            # Save workbook
                            wb.save(xl_filename)
                            wb.close()
                            summary_wb.save('Trade sum_bear.xlsx')

                            # Print time summary
                            time_elapsed = datetime.datetime.now() - start
                            print(f'Iterations: {count} / 7200')
                            print(f'Time elapsed = {time_elapsed}')
                            print(f'Time remaining: {time_elapsed * (7200 / count) - time_elapsed}.')
                            print('----------------------------------------')
                            count +=1

summary_wb.save('Trade Summary - bearish.xlsx')
summary_wb.close()
print('COMPLETE!')
